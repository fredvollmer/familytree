from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import textwrap
import zipfile
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Optional, List, Dict, Tuple

BASE = Path('/mnt/data')
OUT_PREFIX = 'Fredric_Vollmer_Maternal_Family_Tree_Records_First'

# ----------------------------
# Canonical collateral dataset
# ----------------------------
@dataclass
class Person:
    pid: str
    name: str
    branch: str
    tier: str
    relationship_fredric: str
    relationship_jan: str
    status: str = 'Status not researched'
    confidence: str = 'B - Strong'
    spouse: Optional[str] = None
    parents: Optional[str] = None
    children_in_scope: Optional[str] = None
    evidence_public: str = ''
    evidence_family: str = ''
    sources: str = ''
    notes: str = ''

people: List[Person] = [
    Person('P001','Fredric Muller Vollmer','Immediate family','Root','Self','Son of Jan Muller Vollmer','Living','A - Confirmed',parents='Jan Muller Vollmer and Henry Richard “Dick” Vollmer',sources='S1',notes='Living dates intentionally omitted.'),
    Person('P002','Arianna Lynn “Annie” Vollmer','Immediate family','Sibling','Sister','Daughter of Jan Muller Vollmer','Living','A - Confirmed',parents='Jan Muller Vollmer and Henry Richard “Dick” Vollmer',sources='CF1',notes='Name form follows family correspondence; living dates intentionally omitted.'),
    Person('P003','Jan Muller Vollmer','Immediate family','Parent','Mother','Self','Living','A - Confirmed',spouse='Henry Richard “Dick” Vollmer',parents='Arthur John “Art” Muller and Anne Rives McCormick Muller',children_in_scope='Fredric; Arianna “Annie”',sources='S1; S2; CF1'),
    Person('P004','Henry Richard “Dick” Vollmer','Immediate family','Spouse context','Father','Husband','Deceased','A - Confirmed',spouse='Jan Muller Vollmer',children_in_scope='Fredric; Arianna “Annie”',sources='CF1',notes='Included as spouse/parent context; this project does not trace his ancestry.'),
    Person('P005','Bruce Muller','Immediate maternal','Jan sibling','Maternal uncle','Brother','Deceased','A - Confirmed',parents='Arthur John “Art” Muller and Anne Rives McCormick Muller',children_in_scope='No children located in the reviewed family chart or corroborating records',sources='S2; CF2',notes='“No children located” is not proof that he had none.'),
    Person('P006','Arthur John “Art” Muller','Muller / Lee','Grandparent anchor','Maternal grandfather','Father','Deceased','A - Confirmed',spouse='Anne Rives McCormick Muller',parents='Arthur Herman Muller and Elma Anna Lee',children_in_scope='Bruce; Jan',evidence_public='Library of Congress Veterans History Project collection anchors identity and military service.',evidence_family='1998 family chart authored by Art.',sources='S2; S5'),
    Person('P007','Anne Rives McCormick Muller','McCormick / Slaughter','Grandparent anchor','Maternal grandmother','Mother','Deceased','A - Confirmed',spouse='Arthur John “Art” Muller',parents='Charles Ellsworth McCormick II and Eloise A. Slaughter',children_in_scope='Bruce; Jan',sources='S1; S2; CF2'),
    Person('P008','Arthur Herman Muller','Muller / Lee','Great-grandparent anchor','Maternal great-grandfather','Paternal grandfather','Deceased','A - Confirmed',spouse='Elma Anna Lee',children_in_scope='Mary Gene; Jane; Arthur John',sources='S2; S24'),
    Person('P009','Elma Anna Lee Muller','Muller / Lee','Great-grandparent anchor','Maternal great-grandmother','Paternal grandmother','Deceased','A - Confirmed',spouse='Arthur Herman Muller',children_in_scope='Mary Gene; Jane; Arthur John',sources='S2; S3; S24'),
    Person('P010','Mary Gene Muller Chaffee','Muller / Lee','Art sibling','Maternal great-aunt','Paternal aunt','Status not fully researched','A - Confirmed',spouse='James Chaffee',parents='Arthur Herman Muller and Elma Anna Lee',children_in_scope='Linda; Lori; James',sources='S2; CF3',notes='Child set transcribed from the 1998 chart. Earlier draft mistakenly read James as “Janet”; corrected here.'),
    Person('P011','James Chaffee (husband of Mary Gene)','Muller / Lee','Spouse of Art sibling','Maternal great-uncle by marriage','Paternal uncle by marriage','Status not researched','A - Confirmed',spouse='Mary Gene Muller Chaffee',children_in_scope='Linda; Lori; James',sources='S2; CF3'),
    Person('P012','Linda Chaffee','Muller / Lee','Jan first cousin','First cousin once removed','Paternal first cousin','Living/status not researched','A - Confirmed',parents='Mary Gene Muller Chaffee and James Chaffee',sources='S2'),
    Person('P013','Lori Chaffee Mettler','Muller / Lee','Jan first cousin','First cousin once removed','Paternal first cousin','Living','A - Confirmed',parents='Mary Gene Muller Chaffee and James Chaffee',sources='S2; CF3'),
    Person('P014','James Chaffee (son of Mary Gene and James)','Muller / Lee','Jan first cousin','First cousin once removed','Paternal first cousin','Living/status not researched','A - Confirmed',parents='Mary Gene Muller Chaffee and James Chaffee',sources='S2; CF3',notes='Corrected from the earlier erroneous transcription “Janet Chaffee.”'),
    Person('P015','Jane Muller Swick','Muller / Lee','Art sibling','Maternal great-aunt','Paternal aunt','Status not fully researched','A - Confirmed',spouse='Howard Swick',parents='Arthur Herman Muller and Elma Anna Lee',children_in_scope='Kathy; John; Susan; Patti',sources='S2; CF4'),
    Person('P016','Howard Swick','Muller / Lee','Spouse of Art sibling','Maternal great-uncle by marriage','Paternal uncle by marriage','Status not researched','A - Confirmed',spouse='Jane Muller Swick',children_in_scope='Kathy; John; Susan; Patti',sources='S2'),
    Person('P017','Kathy Swick Werblo','Muller / Lee','Jan first cousin','First cousin once removed','Paternal first cousin','Living','A - Confirmed',spouse='Dave Werblo',parents='Jane Muller Swick and Howard Swick',evidence_family='1955 family photograph label and later correspondence explicitly call Kathy and Jan cousins.',sources='S2; CF4'),
    Person('P018','Dave Werblo','Muller / Lee','Spouse of Jan first cousin','Cousin-in-law (1C1R spouse)','Husband of paternal first cousin','Living','A - Confirmed',spouse='Kathy Swick Werblo',sources='CF4'),
    Person('P019','John Swick','Muller / Lee','Jan first cousin','First cousin once removed','Paternal first cousin','Living/status not researched','A - Confirmed',spouse='Celeste',parents='Jane Muller Swick and Howard Swick',sources='S2; CF4'),
    Person('P020','Celeste (surname not independently verified)','Muller / Lee','Spouse of Jan first cousin','Cousin-in-law (1C1R spouse)','Spouse of paternal first cousin','Status not researched','C - Provisional',spouse='John Swick',sources='S2',notes='Only the first name is legible/available in the 1998 chart.'),
    Person('P021','Susan “Sue” Swick Galbraith','Muller / Lee','Jan first cousin','First cousin once removed','Paternal first cousin','Living','A - Confirmed',spouse='Gordon Galbraith',parents='Jane Muller Swick and Howard Swick',evidence_family='Jan explicitly identifies Sue’s children as Fredric’s second cousins.',sources='S2; CF4'),
    Person('P022','Gordon Galbraith','Muller / Lee','Spouse of Jan first cousin','Cousin-in-law (1C1R spouse)','Husband of paternal first cousin','Status not researched','A - Confirmed',spouse='Susan “Sue” Swick Galbraith',sources='S2'),
    Person('P023','Patti Swick (later surname Silva)','Muller / Lee','Jan first cousin','First cousin once removed','Paternal first cousin','Living','A - Confirmed',spouse='Rick Lilly (shown in 1998 chart; later marital status not assumed)',parents='Jane Muller Swick and Howard Swick',sources='S2; CF3',notes='The 2023 reunion email uses “Patti Swick Silva (PJ),” so Rick Lilly is retained only as the spouse shown in 1998.'),
    Person('P024','Rick Lilly','Muller / Lee','Historical spouse of Jan first cousin','Cousin-in-law (historical)','Historical spouse of paternal first cousin','Status not researched','B - Strong',spouse='Patti Swick',sources='S2'),
    Person('P025','Charles Ellsworth McCormick II','McCormick / Slaughter','Great-grandparent anchor','Maternal great-grandfather','Maternal grandfather','Deceased','A - Confirmed',spouse='Eloise A. Slaughter McCormick',children_in_scope='Anne; Philip; Charles; Gwendolyn (“Gwen”)',sources='S1; S25; CF5'),
    Person('P026','Eloise A. Slaughter McCormick','McCormick / Slaughter','Great-grandparent anchor','Maternal great-grandmother','Maternal grandmother','Deceased','A - Confirmed',spouse='Charles Ellsworth McCormick II',children_in_scope='Anne; Philip; Charles; Gwendolyn (“Gwen”)',sources='S1; S8; CF5'),
    Person('P027','Philip Brooks McCormick','McCormick / Slaughter','Anne sibling','Maternal great-uncle','Maternal uncle','Deceased','A - Confirmed',spouse='Elizabeth “Betsy” Ross McCormick',parents='Charles Ellsworth McCormick II and Eloise A. Slaughter',children_in_scope='Sandy; Ross; Tina',evidence_public='Published obituary/memorial and contemporary engagement notice support identity and marriage.',evidence_family='Jan explicitly identifies Phil as Anne’s brother and Sandy, Ross and Tina as his children.',sources='CF6; P1; P2'),
    Person('P028','Elizabeth “Betsy” Ross McCormick','McCormick / Slaughter','Spouse of Anne sibling','Maternal great-aunt by marriage','Maternal aunt by marriage','Deceased/status not fully researched','B - Strong',spouse='Philip Brooks McCormick',children_in_scope='Sandy; Ross; Tina',evidence_public='Contemporary engagement notice identifies Betsy Ross and Philip Brooks McCormick.',sources='P2; CF6'),
    Person('P029','Sandy McCormick Hill','McCormick / Slaughter','Jan first cousin','First cousin once removed','Maternal first cousin','Living','A - Confirmed',spouse='John Hill',parents='Philip Brooks McCormick and Elizabeth “Betsy” Ross',sources='CF6; CF7'),
    Person('P030','John Hill','McCormick / Slaughter','Spouse of Jan first cousin','Cousin-in-law (1C1R spouse)','Husband of maternal first cousin','Living','A - Confirmed',spouse='Sandy McCormick Hill',sources='CF7'),
    Person('P031','Ross McCormick','McCormick / Slaughter','Jan first cousin','First cousin once removed','Maternal first cousin','Living','A - Confirmed',spouse='Carrie McCormick',parents='Philip Brooks McCormick and Elizabeth “Betsy” Ross',sources='CF6; CF7'),
    Person('P032','Carrie McCormick','McCormick / Slaughter','Spouse of Jan first cousin','Cousin-in-law (1C1R spouse)','Wife of maternal first cousin','Living','A - Confirmed',spouse='Ross McCormick',sources='CF7'),
    Person('P033','Tina McCormick','McCormick / Slaughter','Jan first cousin','First cousin once removed','Maternal first cousin','Living','A - Confirmed',parents='Philip Brooks McCormick and Elizabeth “Betsy” Ross',sources='CF6',notes='Some later correspondence uses the surname Weese; the McCormick birth surname is retained here.'),
    Person('P034','Charles McCormick (full name not yet independently verified)','McCormick / Slaughter','Anne sibling','Maternal great-uncle','Maternal uncle','Deceased','A - Confirmed',spouse='Judy McCormick',parents='Charles Ellsworth McCormick II and Eloise A. Slaughter',children_in_scope='No children identified in the reviewed public and family records',sources='CF5; CF8',notes='Do not infer “no children” from “none identified.” Exact full name and vital dates remain an open record problem.'),
    Person('P035','Judy McCormick','McCormick / Slaughter','Spouse of Anne sibling','Maternal great-aunt by marriage','Maternal aunt by marriage','Living/status not researched','A - Confirmed',spouse='Charles McCormick',children_in_scope='No children identified in the reviewed records',sources='CF7; CF8'),
    Person('P036','Gwendolyn “Gwen” McCormick Hull','McCormick / Slaughter','Anne sibling','Maternal great-aunt','Maternal aunt','Living','A - Confirmed',spouse='Jim Hull',parents='Charles Ellsworth McCormick II and Eloise A. Slaughter',children_in_scope='Carey; Mark',evidence_family='Gwen signs family material as “Aunt Gwen”; post-Philip family correspondence identifies her as the remaining original McCormick sibling.',sources='CF5; CF9; CF10',notes='This branch is a core McCormick sibling household and must be displayed coequally with Philip’s and Charles’s branches.'),
    Person('P037','Jim Hull','McCormick / Slaughter','Spouse of Anne sibling','Maternal great-uncle by marriage','Maternal uncle by marriage','Living','A - Confirmed',spouse='Gwendolyn “Gwen” McCormick Hull',children_in_scope='Carey; Mark',sources='CF10; CF11'),
    Person('P038','Carey Zaura','McCormick / Slaughter','Jan first cousin','First cousin once removed','Maternal first cousin','Living','A - Confirmed',parents='Gwendolyn “Gwen” McCormick Hull and Jim Hull',evidence_family='Carey addresses Jim as Dad and explicitly calls herself his “one and only daughter,” asking him to read the message to Mom and Mark.',sources='CF11',notes='Some correspondence uses the email surname Oialucia.'),
    Person('P039','Mark Hull','McCormick / Slaughter','Jan first cousin','First cousin once removed','Maternal first cousin','Living','A - Confirmed',parents='Gwendolyn “Gwen” McCormick Hull and Jim Hull',evidence_family='Gwen and Jim’s family letter is signed “Gwen, Jim and Mark”; Carey’s direct daughter statement distinguishes Mark as the other child in the household.',sources='CF10; CF11'),
]

households = [
    {
        'household':'Arthur Herman Muller + Elma Anna Lee',
        'branch':'Muller / Lee',
        'children':'Mary Gene Muller; Jane Muller; Arthur John “Art” Muller',
        'result':'Complete sibling set used for this bounded extension.',
        'confidence':'A - Confirmed',
        'sources':'S2; S3; S24',
        'open_issue':'Locate original census/obituary copies for a fully record-cited household reconstruction.'
    },
    {
        'household':'Mary Gene Muller + James Chaffee',
        'branch':'Muller / Lee',
        'children':'Linda Chaffee; Lori Chaffee Mettler; James Chaffee',
        'result':'Corrected child set. “James,” not “Janet.”',
        'confidence':'A - Confirmed / chart-led',
        'sources':'S2; CF3',
        'open_issue':'Obtain a public obituary or census household image to corroborate all three children independently.'
    },
    {
        'household':'Jane Muller + Howard Swick',
        'branch':'Muller / Lee',
        'children':'Kathy; John; Susan “Sue”; Patti',
        'result':'Complete child set at the requested generation. Their children are second cousins to Fredric and intentionally excluded.',
        'confidence':'A - Confirmed / chart plus contemporary corroboration',
        'sources':'S2; CF4',
        'open_issue':'Verify Celeste’s surname and Patti’s later marital chronology.'
    },
    {
        'household':'Arthur John “Art” Muller + Anne Rives McCormick',
        'branch':'Immediate maternal',
        'children':'Bruce Muller; Jan Muller Vollmer',
        'result':'Jan has one documented sibling, Bruce. No children of Bruce were located in the reviewed sources.',
        'confidence':'A - Confirmed',
        'sources':'S2; CF2',
        'open_issue':'A negative finding is not proof Bruce had no descendants; obituary/probate records would settle this.'
    },
    {
        'household':'Charles Ellsworth McCormick II + Eloise A. Slaughter',
        'branch':'McCormick / Slaughter',
        'children':'Anne Rives; Philip Brooks; Charles; Gwendolyn “Gwen”',
        'result':'Four-child sibling set. Gwen is a core sibling, not a peripheral correspondent.',
        'confidence':'A - Confirmed',
        'sources':'S1; S8; S25; CF5; CF6; CF9',
        'open_issue':'Locate one public obituary or census source that names all four children in one record.'
    },
    {
        'household':'Philip Brooks McCormick + Elizabeth “Betsy” Ross',
        'branch':'McCormick / Slaughter',
        'children':'Sandy McCormick Hill; Ross McCormick; Tina McCormick',
        'result':'Complete child set at requested generation.',
        'confidence':'A - Confirmed',
        'sources':'CF6; P1; P2',
        'open_issue':'Living descendants beyond this generation intentionally excluded.'
    },
    {
        'household':'Charles McCormick + Judy McCormick',
        'branch':'McCormick / Slaughter',
        'children':'No children identified',
        'result':'Couple included. No child is asserted or excluded without an obituary/probate source.',
        'confidence':'A for couple; unresolved for descendants',
        'sources':'CF7; CF8',
        'open_issue':'Resolve Charles’s full name, vital dates, obituary and any descendants.'
    },
    {
        'household':'Gwendolyn “Gwen” McCormick + Jim Hull',
        'branch':'McCormick / Slaughter',
        'children':'Carey Zaura; Mark Hull',
        'result':'Complete two-child household at requested generation; Carey is the one and only daughter.',
        'confidence':'A - Confirmed',
        'sources':'CF9; CF10; CF11',
        'open_issue':'Living dates omitted by design.'
    },
]

corrections = [
    ('Gwen branch prominence','Earlier edition treated Gwen’s household as secondary/provisional.','Gwen is displayed as one of the four children of Charles II and Eloise, coequal with Anne, Philip and Charles.','Structural correction'),
    ('Carey and Mark parentage','Earlier edition used broad family-email grouping.','Carey’s direct statement that she is Jim’s one and only daughter, plus the Gwen/Jim/Mark family letter, supports Carey and Mark as the two children.','Evidence upgrade'),
    ('Chaffee child name','Earlier edition listed “Janet Chaffee.”','The chart name is James Chaffee. Janet is removed everywhere.','Transcription correction'),
    ('Relationship terminology','Earlier narrative risked calling Jan’s cousins Fredric’s first cousins.','They are Fredric’s first cousins once removed. Fredric has no identified maternal first cousins through Bruce in the reviewed evidence.','Kinship correction'),
    ('Negative evidence','Earlier wording could be read as saying Bruce or Charles had no children.','Now says “no children identified/located,” explicitly distinguishing absence of evidence from evidence of absence.','Method correction'),
    ('Email weighting','Earlier extension discovered and inferred relationships mainly through email frequency and recipient lists.','Households are reconstructed from parent couples downward; the 1998 chart, public records and direct statements outrank recipient-list inference.','Method correction'),
    ('Scope/cardinality','Earlier scope was not explicit enough.','Includes Fredric’s immediate maternal collateral family and Jan’s aunts/uncles and first cousins on both parental branches. Descendants of Jan’s cousins are excluded.','Scope correction'),
]

source_rows = [
    ('CF1','Immediate-family correspondence and family-authored documents','Private family evidence','Anchors Jan, Dick, Fredric and Annie; living dates omitted.'),
    ('CF2','1998 Art Muller chart plus Jan’s recollections of Bruce','Family-authored chart and direct testimony','Supports Art/Anne children Bruce and Jan; chart does not show descendants for Bruce.'),
    ('CF3','1998 chart and 2023 Muller reunion correspondence','Family-authored chart and contemporary correspondence','Supports Mary Gene/James Chaffee household; later network corroborates Lori and Jim/James. Correct chart transcription is James.'),
    ('CF4','Kathy Werblo–Jan Muller correspondence, 1955 photo label, 2022 cousin update','Direct cousin testimony and family image','Independently corroborates Jane/Howard Swick branch and Jan–Kathy cousin relationship.'),
    ('CF5','McCormick sibling-family correspondence and inherited family documents','Direct family evidence','Reconstructs children of Charles II and Eloise as Anne, Philip, Charles and Gwen; used with public records where available.'),
    ('CF6','Jan Muller Vollmer email “Phil” (21 Feb 2021)','Direct family statement','Explicitly states Philip is Anne/Nana’s brother and Sandy, Ross and Tina are his children.'),
    ('CF7','McCormick family correspondence, 2018–2026','Contemporary direct family evidence','Corroborates spouses Judy, John Hill and Carrie and maintains the sibling-family network.'),
    ('CF8','Jan correspondence concerning Charles’s Navy memorial and Judy’s trust','Direct family evidence','Supports Charles/Judy couple; exact full name and descendants remain unresolved.'),
    ('CF9','Gwendolyn Hull message signed “Aunt Gwen” and sibling-family correspondence','Direct family evidence','Supports Gwen’s aunt role and placement in the McCormick sibling generation.'),
    ('CF10','Gwen and Jim Hull 2022 family letter','Family-authored household letter','Signed “Gwen, Jim and Mark” and discusses Carey nearby; household corroboration.'),
    ('CF11','Carey Zaura message to her father, 11 Jul 2025','Direct child-to-parent statement','Carey calls herself his “one and only daughter” and asks him to read it to Mom and Mark.'),
    ('P1','Philip Brooks McCormick obituary/memorial','Public obituary','Public identity/death corroboration for Philip Brooks McCormick.', 'https://www.legacy.com/us/obituaries/latimes/name/philip-mccormick-obituary?id=6809282'),
    ('P2','Betsy Ross–Philip Brooks McCormick engagement notice','Contemporary newspaper notice','Public marriage/parental context for Philip and Betsy.', 'https://www.newspapers.com/article/the-daily-breeze-betsy-ross-and-philip-b/165847296/'),
    ('P3','Arthur John Muller Collection, Veterans History Project','Library of Congress record','Public identity and military-service anchor for Art Muller.', 'https://www.loc.gov/item/afc2001001.95688/'),
]

# ----------------------------
# JSON data artifact
# ----------------------------
json_path = BASE / f'{OUT_PREFIX}_Canonical_Data.json'
json_path.write_text(json.dumps({
    'scope': 'Direct maternal ancestry retained from prior edition; corrected bounded collateral extension through Jan’s aunts/uncles and first cousins on both parental branches, plus spouses and immediate family.',
    'people': [asdict(p) for p in people],
    'households': households,
    'corrections': [dict(zip(['topic','prior','corrected','type'], c)) for c in corrections],
    'sources': [dict(zip(['id','title','type','use','url'], row if len(row)==5 else (*row, ''))) for row in source_rows]
}, indent=2, ensure_ascii=False), encoding='utf-8')

# ----------------------------
# Workbook
# ----------------------------
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

src_xlsx = BASE / 'Fredric_Vollmer_Maternal_Ancestor_Index.xlsx'
wb = load_workbook(src_xlsx)

# Remove/replace any earlier collateral sheets.
for sname in ['Extended Family','Household Audit','Correction Log']:
    if sname in wb.sheetnames:
        del wb[sname]

# Update Read Me scope and add audit metrics.
ws = wb['Read Me']
ws['A1'] = 'Maternal Family Tree of Fredric Muller Vollmer — Records-First Corrected Edition'
ws['A2'] = 'Direct ancestry plus bounded collateral-family audit — 30 August 2026'
ws['A4'] = ('Scope: the 169-person direct maternal ancestor register is retained. The corrected collateral extension reconstructs each '
            'grandparental household from the parents downward, adds Fredric’s immediate maternal family, and includes Jan’s aunts/uncles '
            'and first cousins on both the Muller/Lee and McCormick/Slaughter branches. Children of Jan’s cousins are outside scope. '
            'Living dates and private contact details are omitted.')
ws.merge_cells('A4:E4') if not any(str(rng)=='A4:E4' for rng in ws.merged_cells.ranges) else None
ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')
ws.row_dimensions[4].height = 66
# Add a collateral summary block lower down.
start = 23
summary_values = [
    ('Collateral audit metric','Value'),
    ('People in corrected collateral dataset',len(people)),
    ('Households audited',len(households)),
    ('Jan first cousins identified',12),
    ('Fredric maternal first cousins identified through Jan’s sibling line',0),
    ('Unresolved descendant questions',2),
]
for r, row in enumerate(summary_values, start=start):
    for c, val in enumerate(row, start=1):
        ws.cell(r,c,val)
for cell in ws[start]:
    cell.fill = PatternFill('solid', fgColor='17365D')
    cell.font = Font(color='FFFFFF', bold=True)

# New Extended Family sheet.
ext = wb.create_sheet('Extended Family', 3)
headers = ['ID','Person','Branch','Tier','Relationship to Fredric','Relationship to Jan','Spouse / partner','Parents','Children in bounded scope','Status','Confidence','Public-record evidence','Family-document evidence','Source IDs','Notes']
ext.append(headers)
for p in people:
    ext.append([p.pid,p.name,p.branch,p.tier,p.relationship_fredric,p.relationship_jan,p.spouse,p.parents,p.children_in_scope,p.status,p.confidence,p.evidence_public,p.evidence_family,p.sources,p.notes])

# Household Audit sheet.
ha = wb.create_sheet('Household Audit', 4)
ha_headers = ['Household','Branch','Children at bounded generation','Audit result','Confidence','Source IDs','Remaining record problem']
ha.append(ha_headers)
for h in households:
    ha.append([h['household'],h['branch'],h['children'],h['result'],h['confidence'],h['sources'],h['open_issue']])

# Correction Log sheet.
cl = wb.create_sheet('Correction Log', 5)
cl_headers = ['Topic','Earlier treatment','Corrected treatment','Correction type']
cl.append(cl_headers)
for row in corrections:
    cl.append(list(row))

# Research gaps: append collateral gaps.
rg = wb['Research Gaps']
existing = rg.max_row + 1
new_gaps = [
    ['Collateral: Bruce Muller descendants','No spouse or children located in the reviewed chart/records.','Obituary, probate, cemetery record, family confirmation.','High','Do not convert this negative search result into “no children.”'],
    ['Collateral: Charles McCormick identity/descendants','Full name, vital dates and any children remain unresolved.','Obituary, probate/trust documents, Navy service record, California vital indexes.','High','Couple with Judy is confirmed; descendant question remains open.'],
    ['Collateral: Mary Gene Chaffee household corroboration','1998 chart supplies Linda, Lori and James.','Census/obituary copies for Mary Gene or James Chaffee.','Medium','Earlier “Janet” transcription corrected to James.'],
    ['Collateral: Celeste surname','Only first name available on 1998 chart.','Marriage notice or family confirmation for John Swick.','Low','Keep surname blank until sourced.'],
]
for row in new_gaps:
    rg.append(row)

# Source key: append corrected collateral sources.
sk = wb['Source Key']
for row in source_rows:
    if len(row)==4:
        sid,title,typ,use = row
        url=''
    else:
        sid,title,typ,use,url = row
    sk.append([sid,title,typ,url,use])

# Styling helpers.
header_fill = PatternFill('solid', fgColor='17365D')
sub_fill = PatternFill('solid', fgColor='D9EAF7')
thin = Side(style='thin', color='B7C9D6')
for sheet in [ext,ha,cl]:
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet.row_dimensions[1].height = 34
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = Border(bottom=thin)
    for r in range(2, sheet.max_row+1):
        if r % 2 == 0:
            for cell in sheet[r]:
                cell.fill = PatternFill('solid', fgColor='F4F8FB')

ext_widths = [8,31,22,22,27,25,35,38,36,23,17,48,52,20,58]
for i,w in enumerate(ext_widths,1):
    ext.column_dimensions[get_column_letter(i)].width = w
for r in range(2, ext.max_row+1):
    ext.row_dimensions[r].height = 68
ha_widths = [40,23,45,55,25,22,65]
for i,w in enumerate(ha_widths,1):
    ha.column_dimensions[get_column_letter(i)].width = w
for r in range(2, ha.max_row+1):
    ha.row_dimensions[r].height = 70
cl_widths = [28,57,68,24]
for i,w in enumerate(cl_widths,1):
    cl.column_dimensions[get_column_letter(i)].width = w
for r in range(2, cl.max_row+1):
    cl.row_dimensions[r].height = 62

# Extend source/research widths and wrapping.
for sheet_name in ['Research Gaps','Source Key','Read Me']:
    sh=wb[sheet_name]
    for row in sh.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

# Add tables (unique names).
for sheet, ref, name in [
    (ext, f'A1:O{ext.max_row}', 'CorrectedExtendedFamily'),
    (ha, f'A1:G{ha.max_row}', 'HouseholdAuditTable'),
    (cl, f'A1:D{cl.max_row}', 'CorrectionLogTable')]:
    tab=Table(displayName=name, ref=ref)
    tab.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    sheet.add_table(tab)

xlsx_path = BASE / f'{OUT_PREFIX}_Index.xlsx'
wb.save(xlsx_path)

# ----------------------------
# Graphviz charts
# ----------------------------
DOT = r'''
digraph G {
  graph [rankdir=TB, bgcolor="white", pad="0.25", nodesep="0.30", ranksep="0.50", splines=ortho, fontname="DejaVu Sans"];
  node [shape=box, style="rounded,filled", fontname="DejaVu Sans", fontsize=13, margin="0.13,0.09", color="#2F4F4F", penwidth=1.3];
  edge [color="#566573", penwidth=1.1, arrowsize=0.6];

  legend [label="Solid outline: confirmed/strong  |  Dashed outline: unresolved detail\nLiving dates omitted  |  Children of Jan's cousins excluded", fillcolor="#F7F9F9", fontsize=11];

  ah [label="Arthur Herman Muller\n& Elma Anna Lee", fillcolor="#D9EAF7"];
  ce [label="Charles Ellsworth McCormick II\n& Eloise A. Slaughter", fillcolor="#FCE4D6"];
  mg [label="Mary Gene Muller\n& James Chaffee", fillcolor="#EAF2F8"];
  js [label="Jane Muller\n& Howard Swick", fillcolor="#EAF2F8"];
  aa [label="Arthur John 'Art' Muller\n& Anne Rives McCormick", fillcolor="#E2F0D9", penwidth=2.0];
  pm [label="Philip Brooks McCormick\n& Elizabeth 'Betsy' Ross", fillcolor="#FDE9D9"];
  cm [label="Charles McCormick\n& Judy McCormick", fillcolor="#FDE9D9", style="rounded,filled,dashed"];
  gh [label="Gwendolyn 'Gwen' McCormick\n& Jim Hull", fillcolor="#FDE9D9", penwidth=2.0];

  ah -> mg; ah -> js; ah -> aa;
  ce -> aa; ce -> pm; ce -> cm; ce -> gh;

  mgc [label="Linda Chaffee\nLori Chaffee Mettler\nJames Chaffee", fillcolor="#F8FBFD"];
  jsc [label="Kathy Swick Werblo\nJohn Swick\nSusan 'Sue' Swick Galbraith\nPatti Swick", fillcolor="#F8FBFD"];
  aac [label="Bruce Muller\nJan Muller Vollmer", fillcolor="#F2F8ED"];
  pmc [label="Sandy McCormick Hill\nRoss McCormick\nTina McCormick", fillcolor="#FFF8F2"];
  cmc [label="No children identified\n(record question remains open)", fillcolor="#FFF8F2", style="rounded,filled,dashed", fontsize=11];
  ghc [label="Carey Zaura\nMark Hull", fillcolor="#FFF8F2", penwidth=2.0];

  mg -> mgc; js -> jsc; aa -> aac; pm -> pmc; cm -> cmc; gh -> ghc;

  jv [label="Jan Muller Vollmer\n& Henry Richard 'Dick' Vollmer", fillcolor="#DDEBF7", penwidth=2.0];
  roots [label="Fredric Muller Vollmer\nArianna Lynn 'Annie' Vollmer", fillcolor="#E2F0D9", penwidth=2.0];
  aac -> jv [label=" Jan ", fontsize=10];
  jv -> roots;

  {rank=same; ah; ce}
  {rank=same; mg; js; aa; pm; cm; gh}
  {rank=same; mgc; jsc; aac; pmc; cmc; ghc}
}
'''
dot_path = BASE / f'{OUT_PREFIX}_Chart.dot'
dot_path.write_text(DOT, encoding='utf-8')
svg_path = BASE / f'{OUT_PREFIX}_Chart.svg'
png_path = BASE / f'{OUT_PREFIX}_Chart.png'
try:
    subprocess.run(['dot','-Tsvg',str(dot_path),'-o',str(svg_path)],check=True)
    subprocess.run(['dot','-Tpng','-Gdpi=180',str(dot_path),'-o',str(png_path)],check=True)
except Exception:
    # Fallback: preserve DOT if Graphviz unavailable.
    svg_path.write_text('<svg xmlns="http://www.w3.org/2000/svg" width="1200" height="300"><text x="20" y="40" font-size="24">See DOT source for corrected extended family chart.</text></svg>',encoding='utf-8')

# ----------------------------
# GEDCOM extension
# ----------------------------
base_ged = BASE / 'Fredric_Vollmer_Maternal_Family_Tree.ged'
ged_text = base_ged.read_text(encoding='utf-8', errors='replace')
# Strip trailer so records may be appended.
ged_body = re.sub(r'\n0 TRLR\s*$', '\n', ged_text.strip()) + '\n'

# Parse existing individual names.
existing_ids: Dict[str,str] = {}
current=None
for line in ged_body.splitlines():
    m=re.match(r'0 (@I\d+@) INDI',line)
    if m:
        current=m.group(1)
    elif current and line.startswith('1 NAME '):
        nm=line[7:].replace('/','').strip().lower()
        existing_ids[nm]=current

max_i=max([int(x.strip('@I')) for x in re.findall(r'0 (@I\d+@) INDI',ged_body)] or [0])
max_f=max([int(x.strip('@F')) for x in re.findall(r'0 (@F\d+@) FAM',ged_body)] or [0])

def find_existing(*needles: str) -> Optional[str]:
    needles=[n.lower() for n in needles]
    for nm,iid in existing_ids.items():
        if all(n in nm for n in needles):
            return iid
    return None

def ged_name(name: str) -> str:
    # Treat final word as surname unless known compound/parenthetical. Keep readable; GEDCOM importers tolerate this.
    clean=re.sub(r'\s*\([^)]*\)','',name).replace('“','').replace('”','').replace('’',"'")
    parts=clean.split()
    if not parts: return clean
    surname=parts[-1]
    given=' '.join(parts[:-1])
    return f'{given} /{surname}/'

# Existing anchors.
anchor = {
    'fredric': find_existing('fredric','vollmer'),
    'jan': find_existing('jan','muller','vollmer') or find_existing('jan','vollmer'),
    'art': find_existing('arthur','muller'),
    'anne': find_existing('anne','mccormick') or find_existing('anne','muller'),
    'charles2': find_existing('charles','ellsworth','mccormick'),
    'eloise': find_existing('eloise','mccormick') or find_existing('eloise','slaughter'),
    'arthurherman': find_existing('arthur','herman','muller'),
    'elma': find_existing('elma','lee') or find_existing('elma','muller'),
}

# Add only people not already in direct GEDCOM. Use stable canonical pid map.
new_records=[]
pid_to_iid: Dict[str,str]={}
for key,val in anchor.items():
    if val:
        pid_to_iid[key]=val

# Explicit mapping from canonical person IDs to existing anchors.
existing_pid_map={'P001':'fredric','P003':'jan','P006':'art','P007':'anne','P008':'arthurherman','P009':'elma','P025':'charles2','P026':'eloise'}
for pid,key in existing_pid_map.items():
    if anchor.get(key):
        pid_to_iid[pid]=anchor[key]

for p in people:
    if p.pid in pid_to_iid:
        continue
    max_i += 1
    iid=f'@I{max_i}@'
    pid_to_iid[p.pid]=iid
    sex='F' if any(x in p.name.lower() for x in ['arianna','jan ','anne ','elma','mary gene','linda','lori','jane ','kathy','celeste','susan','patti','betsy','sandy','carrie','tina','judy','gwendolyn','carey']) else 'M'
    rec=[f'0 {iid} INDI',f'1 NAME {ged_name(p.name)}',f'1 SEX {sex}',f'1 NOTE Relationship to Fredric: {p.relationship_fredric}.',f'1 NOTE Confidence: {p.confidence}.']
    if p.status=='Deceased':
        rec.append('1 DEAT')
    new_records.extend(rec)

# Build families by person IDs.
name_to_pid={p.name:p.pid for p in people}
def pid(name): return name_to_pid.get(name)

def add_family(husb_pid: Optional[str], wife_pid: Optional[str], child_pids: List[str], note: str=''):
    global max_f
    nonlocal_dummy=None
    fam=[]
    # Find actual ids.
    h=pid_to_iid.get(husb_pid) if husb_pid else None
    w=pid_to_iid.get(wife_pid) if wife_pid else None
    cs=[pid_to_iid[c] for c in child_pids if c in pid_to_iid]
    if not (h or w) or not cs:
        return
    global_records=[]
    globals()['max_f'] += 1
    fid=f'@F{globals()["max_f"]}@'
    global_records.append(f'0 {fid} FAM')
    if h: global_records.append(f'1 HUSB {h}')
    if w: global_records.append(f'1 WIFE {w}')
    for c in cs: global_records.append(f'1 CHIL {c}')
    if note: global_records.append(f'1 NOTE {note}')
    new_records.extend(global_records)

# Determine sex ordering pragmatically.
add_family('P008','P009',['P010','P015','P006'],'Children reconstructed from the 1998 Art Muller family chart.')
add_family('P011','P010',['P012','P013','P014'],'James is the corrected third child name; earlier draft said Janet.')
add_family('P016','P015',['P017','P019','P021','P023'])
add_family('P006','P007',['P005','P003'])
add_family('P004','P003',['P001','P002'])
add_family('P025','P026',['P007','P027','P034','P036'],'Four-child McCormick sibling set; Gwen is a core sibling.')
add_family('P027','P028',['P029','P031','P033'])
add_family('P037','P036',['P038','P039'],'Carey is the one and only daughter; Mark is the other child.')

# Spouse-only families without children are intentionally not added unless they clarify couple links.
def add_spouse_family(a_pid,b_pid,note=''):
    globals()['max_f'] += 1
    fid=f'@F{globals()["max_f"]}@'
    a=pid_to_iid.get(a_pid); b=pid_to_iid.get(b_pid)
    if not a or not b: return
    # Assume first male when possible.
    rec=[f'0 {fid} FAM',f'1 HUSB {a}',f'1 WIFE {b}']
    if note: rec.append(f'1 NOTE {note}')
    new_records.extend(rec)

for a,b,n in [
    ('P018','P017','Dave and Kathy Werblo'),('P019','P020','John and Celeste'),('P022','P021','Gordon and Susan Galbraith'),
    ('P024','P023','Historical spouse shown in 1998 chart'),('P030','P029','John and Sandy Hill'),('P031','P032','Ross and Carrie McCormick'),('P034','P035','Charles and Judy McCormick')]:
    add_spouse_family(a,b,n)

ged_out = ged_body + '\n'.join(new_records) + '\n0 TRLR\n'
ged_path = BASE / f'{OUT_PREFIX}.ged'
ged_path.write_text(ged_out, encoding='utf-8')

# ----------------------------
# PDF front section and merge
# ----------------------------
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle,
                                Image as RLImage, KeepTogether, LongTable)
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

font_regular='/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'
font_bold='/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'
if Path(font_regular).exists():
    pdfmetrics.registerFont(TTFont('DVSans',font_regular))
    pdfmetrics.registerFont(TTFont('DVSans-Bold',font_bold))
    BODY_FONT='DVSans'; BOLD_FONT='DVSans-Bold'
else:
    BODY_FONT='Helvetica'; BOLD_FONT='Helvetica-Bold'

styles=getSampleStyleSheet()
styles.add(ParagraphStyle(name='FVTitle',parent=styles['Title'],fontName=BOLD_FONT,fontSize=23,leading=28,textColor=colors.HexColor('#17365D'),alignment=TA_CENTER,spaceAfter=16))
styles.add(ParagraphStyle(name='FVSub',parent=styles['Normal'],fontName=BODY_FONT,fontSize=13,leading=18,textColor=colors.HexColor('#334E68'),alignment=TA_CENTER,spaceAfter=12))
styles.add(ParagraphStyle(name='FVH1',parent=styles['Heading1'],fontName=BOLD_FONT,fontSize=18,leading=22,textColor=colors.HexColor('#17365D'),spaceBefore=8,spaceAfter=10))
styles.add(ParagraphStyle(name='FVH2',parent=styles['Heading2'],fontName=BOLD_FONT,fontSize=14.5,leading=18,textColor=colors.HexColor('#8A3B12'),spaceBefore=8,spaceAfter=7))
styles.add(ParagraphStyle(name='FVBody',parent=styles['BodyText'],fontName=BODY_FONT,fontSize=11.2,leading=15.2,spaceAfter=7))
styles.add(ParagraphStyle(name='FVSmall',parent=styles['BodyText'],fontName=BODY_FONT,fontSize=9.4,leading=12.3,spaceAfter=4))
styles.add(ParagraphStyle(name='FVNote',parent=styles['BodyText'],fontName=BODY_FONT,fontSize=10.5,leading=14,backColor=colors.HexColor('#F4F8FB'),borderColor=colors.HexColor('#A9C4D8'),borderWidth=0.8,borderPadding=8,spaceBefore=6,spaceAfter=8))
styles.add(ParagraphStyle(name='FVTable',parent=styles['BodyText'],fontName=BODY_FONT,fontSize=8.7,leading=11.2))
styles.add(ParagraphStyle(name='FVTableHead',parent=styles['BodyText'],fontName=BOLD_FONT,fontSize=8.8,leading=11,textColor=colors.white,alignment=TA_CENTER))

front_pdf = BASE / f'{OUT_PREFIX}_Front_Section.pdf'

def footer(canvas, doc):
    canvas.saveState()
    canvas.setFont(BODY_FONT,8)
    canvas.setFillColor(colors.HexColor('#5D6D7E'))
    canvas.drawString(0.65*inch,0.42*inch,'Fredric Vollmer maternal family tree — records-first corrected edition')
    canvas.drawRightString(7.85*inch,0.42*inch,f'Page {doc.page}')
    canvas.restoreState()

story=[]
story += [Spacer(1,0.7*inch), Paragraph('Maternal Family Tree of<br/>Fredric Muller Vollmer',styles['FVTitle']),
          Paragraph('Records-First Corrected Extended Edition',styles['FVSub']), Spacer(1,0.22*inch),
          Paragraph('Direct ancestry retained • Collateral households rebuilt from each parent couple downward • Gwen McCormick Hull restored to her proper place in the core McCormick sibling family',styles['FVSub']),
          Spacer(1,0.45*inch),
          Paragraph('<b>Research date:</b> 30 August 2026',styles['FVBody']),
          Paragraph('<b>Privacy:</b> dates, addresses, telephone numbers and email addresses for living people are intentionally omitted.',styles['FVBody']),
          Spacer(1,0.35*inch),
          Paragraph('This edition supersedes the collateral-family section of the earlier “Extended” report. The 169-person direct-ancestor register and the Robert E. Lee research are retained as an appendix.',styles['FVNote']),
          PageBreak()]

story += [Paragraph('Executive findings',styles['FVH1']),
          Paragraph('The earlier collateral extension had a methodological weakness: it discovered family structure largely through email threads. That favored frequent correspondents and made Gwen’s branch look peripheral. This edition starts with the two great-grandparent couples, enumerates each child, then reconstructs every child’s bounded household.',styles['FVBody']),
          Paragraph('<b>Gwendolyn “Gwen” McCormick Hull is one of the four children of Charles Ellsworth McCormick II and Eloise Slaughter.</b> She is Anne’s sister, Jan’s maternal aunt, and Fredric’s maternal great-aunt. Her household with Jim Hull and children Carey and Mark is displayed at the same structural level as Philip’s and Charles’s households.',styles['FVNote']),
          Paragraph('<b>Immediate maternal collateral result:</b> Jan has one documented sibling, Bruce Muller. No spouse or children of Bruce were located in the reviewed sources. Therefore, this audit identifies no maternal first cousins of Fredric through Jan’s sibling line—but does not claim that none ever existed.',styles['FVBody']),
          Paragraph('<b>Broader two-branch result:</b> Jan has twelve identified first cousins across the Muller and McCormick branches. They are Fredric’s first cousins once removed.',styles['FVBody']),
          Paragraph('<b>Cardinality:</b> the bounded extension is not a cardinality problem. It contains 39 people including spouses and anchor ancestors. Expanding one more generation to every child of Jan’s twelve cousins would substantially enlarge the tree and remains outside this edition.',styles['FVBody']),
          PageBreak()]

story += [Paragraph('Method and evidence hierarchy',styles['FVH1'])]
method_rows=[
    [Paragraph('Priority',styles['FVTableHead']),Paragraph('Evidence class',styles['FVTableHead']),Paragraph('How it is used',styles['FVTableHead'])],
    [Paragraph('1',styles['FVTable']),Paragraph('Civil, census, military, obituary and contemporary newspaper records',styles['FVTable']),Paragraph('Anchor identity, marriage, dates and deceased household members.',styles['FVTable'])],
    [Paragraph('2',styles['FVTable']),Paragraph('Family-authored documents created for genealogy',styles['FVTable']),Paragraph('The 1998 Art Muller chart is the principal topology source for the Muller collateral families, checked against later evidence.',styles['FVTable'])],
    [Paragraph('3',styles['FVTable']),Paragraph('Direct statements by a close relative',styles['FVTable']),Paragraph('Used where public records do not name living descendants—for example, Jan’s direct statement about Phil and his children and Carey’s statement to her father.',styles['FVTable'])],
    [Paragraph('4',styles['FVTable']),Paragraph('Email recipient lists, surnames and household context',styles['FVTable']),Paragraph('Corroboration only. They are no longer treated as sufficient proof of parentage by themselves.',styles['FVTable'])],
]
t=Table(method_rows,colWidths=[0.65*inch,2.45*inch,4.45*inch],repeatRows=1)
t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#17365D')),('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#A9C4D8')),('VALIGN',(0,0),(-1,-1),'TOP'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F4F8FB')]),('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6)]))
story += [t,Spacer(1,0.12*inch),Paragraph('Confidence grades retain the original report’s definitions: A confirmed; B strong; C provisional; D traditional/disputed. A grade applies to the stated relationship, not automatically to every date or spelling associated with the person.',styles['FVBody']),PageBreak()]

# Chart page.
story += [Paragraph('Corrected extended-family structure',styles['FVH1'])]
if png_path.exists():
    story.append(RLImage(str(png_path),width=7.3*inch,height=5.55*inch))
story += [Spacer(1,0.12*inch),Paragraph('The separate SVG in the package is the preferred zoomable version. Gwen’s household is intentionally drawn with the same weight as the other McCormick sibling households.',styles['FVSmall']),PageBreak()]

# Immediate family.
story += [Paragraph('Immediate maternal family',styles['FVH1']),
          Paragraph('<b>Arthur John “Art” Muller + Anne Rives McCormick Muller</b>',styles['FVH2']),
          Paragraph('Children: Bruce Muller and Jan Muller Vollmer.',styles['FVBody']),
          Paragraph('<b>Jan Muller Vollmer + Henry Richard “Dick” Vollmer</b>',styles['FVH2']),
          Paragraph('Children included in this family context: Fredric Muller Vollmer and Arianna Lynn “Annie” Vollmer.',styles['FVBody']),
          Paragraph('<b>Bruce Muller</b>',styles['FVH2']),
          Paragraph('Bruce is Fredric’s maternal uncle. The 1998 chart and reviewed collateral evidence do not identify a spouse or children for him. The tree records this as an unresolved negative search result, not as a claim of lifelong childlessness.',styles['FVBody']),
          PageBreak()]

# Branch tables function.
def branch_table(branch_name: str, title: str):
    rows=[[Paragraph('Person',styles['FVTableHead']),Paragraph('Relationship to Fredric',styles['FVTableHead']),Paragraph('Household / evidence note',styles['FVTableHead']),Paragraph('Grade',styles['FVTableHead'])]]
    for p in people:
        if p.branch != branch_name: continue
        if p.tier in ['Great-grandparent anchor']: continue
        note=[]
        if p.parents: note.append('Parents: '+p.parents+'.')
        if p.spouse: note.append('Spouse/context: '+p.spouse+'.')
        if p.children_in_scope: note.append('Children in scope: '+p.children_in_scope+'.')
        if p.notes: note.append(p.notes)
        rows.append([Paragraph(p.name,styles['FVTable']),Paragraph(p.relationship_fredric,styles['FVTable']),Paragraph(' '.join(note),styles['FVTable']),Paragraph(p.confidence.split(' - ')[0],styles['FVTable'])])
    tab=LongTable(rows,colWidths=[1.55*inch,1.55*inch,3.95*inch,0.45*inch],repeatRows=1)
    tab.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#17365D')),('GRID',(0,0),(-1,-1),0.45,colors.HexColor('#B7C9D6')),('VALIGN',(0,0),(-1,-1),'TOP'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F7FAFC')]),('LEFTPADDING',(0,0),(-1,-1),4),('RIGHTPADDING',(0,0),(-1,-1),4),('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
    return [Paragraph(title,styles['FVH1']),tab,PageBreak()]

story += branch_table('Muller / Lee','Muller / Lee collateral register')
story += branch_table('McCormick / Slaughter','McCormick / Slaughter collateral register')

# Household audit.
story += [Paragraph('Household-by-household completeness audit',styles['FVH1'])]
for h in households:
    story += [Paragraph(h['household'],styles['FVH2']),
              Paragraph(f"<b>Children in bounded scope:</b> {h['children']}",styles['FVBody']),
              Paragraph(f"<b>Finding:</b> {h['result']}",styles['FVBody']),
              Paragraph(f"<b>Confidence:</b> {h['confidence']} &nbsp;&nbsp; <b>Sources:</b> {h['sources']}",styles['FVSmall']),
              Paragraph(f"<b>Remaining record problem:</b> {h['open_issue']}",styles['FVNote'])]
story.append(PageBreak())

# Corrections.
story += [Paragraph('Corrections to the earlier extended edition',styles['FVH1'])]
corr_rows=[[Paragraph('Issue',styles['FVTableHead']),Paragraph('Earlier treatment',styles['FVTableHead']),Paragraph('Corrected treatment',styles['FVTableHead'])]]
for topic,prior,corr,typ in corrections:
    corr_rows.append([Paragraph(topic,styles['FVTable']),Paragraph(prior,styles['FVTable']),Paragraph(corr,styles['FVTable'])])
ct=LongTable(corr_rows,colWidths=[1.45*inch,2.65*inch,3.45*inch],repeatRows=1)
ct.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#8A3B12')),('GRID',(0,0),(-1,-1),0.45,colors.HexColor('#D8B4A0')),('VALIGN',(0,0),(-1,-1),'TOP'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#FFF8F2')]),('LEFTPADDING',(0,0),(-1,-1),5),('RIGHTPADDING',(0,0),(-1,-1),5),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
story += [ct,PageBreak()]

# Research gaps.
story += [Paragraph('Remaining high-value record work',styles['FVH1']),
          Paragraph('The bounded tree is structurally complete to the requested tier, but “complete” does not mean every detail is proven. These are the remaining questions most likely to alter the collateral tree:',styles['FVBody'])]
for title, desc in [
    ('Bruce Muller obituary/probate','Would determine whether a spouse or children existed and whether Fredric has maternal first cousins through Bruce.'),
    ('Charles McCormick obituary/probate','Would settle his full name, dates and any descendants with Judy.'),
    ('Mary Gene Chaffee public household record','Would independently corroborate Linda, Lori and James beyond the 1998 chart.'),
    ('John Swick marriage record','Would establish Celeste’s full name.'),
]:
    story += [Paragraph(title,styles['FVH2']),Paragraph(desc,styles['FVBody'])]
story += [Paragraph('Children of Jan’s twelve first cousins are intentionally not added. They are Fredric’s second cousins, and adding that generation should be treated as a separate bounded project rather than allowed to grow implicitly.',styles['FVNote']),PageBreak()]

# Source ledger.
story += [Paragraph('Collateral source ledger',styles['FVH1'])]
source_table=[[Paragraph('ID',styles['FVTableHead']),Paragraph('Source',styles['FVTableHead']),Paragraph('Use and caution',styles['FVTableHead'])]]
for row in source_rows:
    sid,title,typ,use = row[:4]
    url=row[4] if len(row)>4 else ''
    label=f'<b>{title}</b><br/><font size="7.8">{typ}</font>'
    use2=use + (f'<br/><font size="7.5">{url}</font>' if url else '')
    source_table.append([Paragraph(sid,styles['FVTable']),Paragraph(label,styles['FVTable']),Paragraph(use2,styles['FVTable'])])
st=LongTable(source_table,colWidths=[0.55*inch,2.8*inch,4.2*inch],repeatRows=1)
st.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#17365D')),('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#B7C9D6')),('VALIGN',(0,0),(-1,-1),'TOP'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F7FAFC')]),('LEFTPADDING',(0,0),(-1,-1),4),('RIGHTPADDING',(0,0),(-1,-1),4),('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
story += [st,PageBreak(),Paragraph('Appendix A — direct ancestry and Robert E. Lee inquiry',styles['FVH1']),Paragraph('The following pages reproduce the prior revised direct-ancestry report. Its direct-ancestor register and Robert E. Lee addendum remain separate from—and are not changed by—the corrected collateral-family audit above.',styles['FVBody'])]

doc=SimpleDocTemplate(str(front_pdf),pagesize=letter,rightMargin=0.6*inch,leftMargin=0.6*inch,topMargin=0.58*inch,bottomMargin=0.62*inch,title='Maternal Family Tree of Fredric Muller Vollmer — Records-First Corrected Extended Edition',author='OpenAI research assistance for Fredric Muller Vollmer')
doc.build(story,onFirstPage=footer,onLaterPages=footer)

# Merge with prior revised direct report.
import fitz
base_report = BASE / 'Fredric_Vollmer_Maternal_Family_Tree_Report_Revised.pdf'
final_pdf = BASE / f'{OUT_PREFIX}_Report.pdf'
outdoc=fitz.open()
outdoc.insert_pdf(fitz.open(str(front_pdf)))
if base_report.exists():
    outdoc.insert_pdf(fitz.open(str(base_report)))
outdoc.set_metadata({'title':'Maternal Family Tree of Fredric Muller Vollmer — Records-First Corrected Extended Edition','author':'OpenAI research assistance for Fredric Muller Vollmer','subject':'Maternal direct ancestry and corrected collateral family tree'})
outdoc.save(str(final_pdf),garbage=4,deflate=True)
outdoc.close()

# ----------------------------
# README and package
# ----------------------------
readme = BASE / f'{OUT_PREFIX}_README.txt'
readme.write_text(f'''MATERNAL FAMILY TREE OF FREDRIC MULLER VOLLMER\nRECORDS-FIRST CORRECTED EXTENDED EDITION\n\nScope\n-----\nThe prior 169-person direct maternal ancestor register is retained. The corrected collateral extension includes Fredric's immediate maternal family and Jan Muller Vollmer's aunts/uncles and first cousins on both her Muller/Lee and McCormick/Slaughter sides, plus spouses needed for context. Children of Jan's cousins are excluded.\n\nPrincipal corrections\n---------------------\n1. Gwendolyn “Gwen” McCormick Hull is restored as a core child of Charles Ellsworth McCormick II and Eloise Slaughter, alongside Anne, Philip and Charles.\n2. Carey Zaura and Mark Hull are recorded as Gwen and Jim Hull's children, supported by a direct child-to-parent statement and a family household letter.\n3. The Chaffee child is James, not Janet.\n4. Jan's first cousins are Fredric's first cousins once removed.\n5. “No children identified” is kept distinct from “had no children.”\n6. Email recipient lists are corroboration only, not standalone proof.\n\nFiles\n-----\n{final_pdf.name} - complete corrected report, followed by prior direct-ancestry appendix\n{xlsx_path.name} - sortable ancestor and collateral registers, household audit, correction log, research gaps and sources\n{ged_path.name} - GEDCOM combining the direct tree with corrected collateral households\n{svg_path.name} - zoomable corrected extended-family chart\n{png_path.name} - PNG version of the chart\n{json_path.name} - canonical machine-readable dataset used to generate the corrected artifacts\n\nPrivacy\n-------\nLiving dates, addresses, telephone numbers and email addresses are omitted.\n''',encoding='utf-8')

zip_path = BASE / f'{OUT_PREFIX}_Package.zip'
with zipfile.ZipFile(zip_path,'w',compression=zipfile.ZIP_DEFLATED) as z:
    for path in [final_pdf,xlsx_path,ged_path,svg_path,png_path,json_path,readme,dot_path]:
        if path.exists(): z.write(path,arcname=path.name)
    for extra in [BASE/'Fredric_Vollmer_Maternal_Family_Tree.svg',BASE/'Fredric_Vollmer_Maternal_Family_Tree_Preview.png',BASE/'MullerFamilyTree.jpeg']:
        if extra.exists(): z.write(extra,arcname='reference/'+extra.name)

# ----------------------------
# Validation
# ----------------------------
validation=[]
# Workbook open check.
wb2=load_workbook(xlsx_path,read_only=True,data_only=False)
validation.append(f'Workbook sheets: {", ".join(wb2.sheetnames)}')
validation.append(f'Extended Family rows: {wb2["Extended Family"].max_row-1}')
validation.append(f'Household Audit rows: {wb2["Household Audit"].max_row-1}')
wb2.close()
# PDF page count and text checks.
doc=fitz.open(str(final_pdf))
validation.append(f'PDF pages: {doc.page_count}')
alltext='\n'.join(page.get_text() for page in doc)
for bad in ['Janet Chaffee']:
    validation.append(f'Forbidden text {bad!r}: {bad in alltext}')
for required in ['Gwendolyn “Gwen” McCormick Hull','James Chaffee','one and only daughter']:
    validation.append(f'Required text {required!r}: {required in alltext}')
doc.close()
# GEDCOM checks.
gt=ged_path.read_text(encoding='utf-8')
validation.append(f'GEDCOM individuals: {len(re.findall(r"^0 @I\\d+@ INDI$",gt,re.M))}')
validation.append(f'GEDCOM families: {len(re.findall(r"^0 @F\\d+@ FAM$",gt,re.M))}')
validation.append(f'GEDCOM trailer count: {gt.count("0 TRLR")}')
# Package listing.
with zipfile.ZipFile(zip_path) as z:
    validation.append(f'Package entries: {len(z.namelist())}')

val_path=BASE/f'{OUT_PREFIX}_VALIDATION.txt'
val_path.write_text('\n'.join(validation)+'\n',encoding='utf-8')
# Add validation to zip after initial close.
with zipfile.ZipFile(zip_path,'a',compression=zipfile.ZIP_DEFLATED) as z:
    z.write(val_path,arcname=val_path.name)

print('\n'.join(validation))
print('OUTPUTS')
for p in [final_pdf,xlsx_path,ged_path,svg_path,png_path,json_path,readme,zip_path,val_path]:
    print(p, p.stat().st_size if p.exists() else 'MISSING')
